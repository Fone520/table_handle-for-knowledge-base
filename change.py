#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import sys
# 拆分所有的合并单元格，并赋予合并之前的值。
# 由于openpyxl并没有提供拆分并填充的方法，所以使用该方法进行完成
def unmerge_and_fill_cells(worksheet):
    all_merged_cell_ranges = list(
        worksheet.merged_cells.ranges
    )
    for merged_cell_range in all_merged_cell_ranges:
        merged_cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)
        for row_index, col_index in merged_cell_range.cells:
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value
# 读取原始xlsx文件，拆分并填充单元格，然后生成中间临时文件。
def unmerge_cell(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        unmerge_and_fill_cells(sheet)
    filename = filename.replace(".xls", "_temp.xls")
    wb.save(filename)
    wb.close()
    return filename
if __name__ == '__main__':
    file = sys.argv[1]
    filename = unmerge_cell(file)


# In[2]:


import pandas as pd
import re
import os

all_sheets = pd.read_excel(filename, sheet_name=None)

def fix(content):
    #print(type(content))
    if type(content) is str:
        content = content.replace('\n', ' ')
        content = content.replace(' ', '')
    return content

if not os.path.exists('temp'):
    os.mkdir('temp')

sheets = []

for sheet_name, df in all_sheets.items():
    sheet_name = sheet_name.replace(' ', '')
    df = df.map(fix)
    with open(f'temp/{sheet_name}.txt','w',encoding='utf8') as f:
        strings = df.to_string(index=False)
        strings = re.sub(r'[ \t]+', ' ', strings)
        sheets.append(strings)
        f.write(strings)
        f.close()
    print(sheet_name)

df = pd.DataFrame({'分段内容':sheets})
df.to_csv('temp/result.csv', index=False)
print('导出csv成功')



